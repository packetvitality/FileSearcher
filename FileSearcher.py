#Built in libraries
import os
import glob
from pathlib import Path
import re
from shutil import move
from shutil import copy
from time import sleep
from time import time
import tarfile
from sys import getfilesystemencoding
import json

#Third party libraries
import magic
from tqdm import tqdm
import openpyxl
import xlrd
from xlrd.sheet import ctype_text
import docx
import PyPDF2

class FileSearcher:
    """
    Provides methods for organizing and searching through files. 
    """
    def __init__(self, working_dir, original_dir, keywords_file, estimated_files=None):
        #Provided to the class
        self.working_dir = working_dir # Base directory we are working from
        self.original_dir = original_dir # Directory containing the files we intend to work with
        self.keywords_file = keywords_file # File containing the keywords which will be searched for
        self.keywords = set()
        with open(self.keywords_file) as file:
            for line in file:
                self.keywords.add(line.strip())
        self.system_encoding = getfilesystemencoding()
        self.estimated_files = estimated_files # Used for more accurately displaying progress with TQDM

    def create_dirs(self):
        """
        Creates the required directories for all functions in the class.
        """
        self.output_dir = os.path.join(self.working_dir, "Output")
        self.processed_dir = os.path.join(self.output_dir, "Processed_Files") # Directory to place files which have been processed
        self.results_dir = os.path.join(self.output_dir, "Results") # Directory to place results
        self.error_dir = os.path.join(self.output_dir, "Error_Files")
        self.unsupported_dir = os.path.join(self.output_dir, "Unsupported_Files")
        self.log_file = os.path.join(self.output_dir, "log.txt") # Used for error logging
        if not os.path.exists(self.output_dir):
            os.mkdir(self.output_dir)
        if not os.path.exists(self.processed_dir):
            os.mkdir(self.processed_dir) 
        if not os.path.exists(self.results_dir):
            os.mkdir(self.results_dir) 
        if not os.path.exists(self.error_dir):
            os.mkdir(self.error_dir)
        if not os.path.exists(self.unsupported_dir):
            os.mkdir(self.unsupported_dir)
        
        # ##  For testing ##
        # self.results_dir = os.path.join(self.working_dir, "temp_Results")
        # self.log_file = os.path.join(self.working_dir, "temp_log.txt")
        # self.processed_dir = os.path.join(self.working_dir, "temp_Search_Complete") # Directory to place files which have already been searched
        # self.error_dir = os.path.join(self.working_dir, "temp_Error_Files")
        # ##              ##      

    def rename_file(self, filename):
        """
        Renames file using 'safe' characters.
        This is useful when operations need to be performed on files but fail due to special characters in the filename.
        Returns either a new file name or a boolean of False
        """
        file_dir = os.path.split(filename)[0]
        file_basename = os.path.split(filename)[1]
        safe_characters = "([^\\w\\.\\-_]|^[\\W_])" #Negate these character sets. In other words, everything but these will be replaced later. 
        # Regex Notes
        # [^\w\.\-_] # negate all of the characters in the set. In other words, these characters are allowed
        # | # or
        # ^[\W_] # negate if it starts with any of the characters in the set. In other words, cannot start with these characters
        # extra \\'s necessary in Python to escape characters
        clean_file_basename = re.sub(safe_characters, "", file_basename)
        new_filename = os.path.join(file_dir, clean_file_basename)
        
        if re.search(safe_characters, file_basename):
            if (len(filename) > 260):
                # Handling long filepaths
                # Reference: https://bugs.python.org/issue18199
                filename = "\\\\?\\" + filename
            try:
                os.rename(filename, new_filename)
                with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                    log_file.write("[{} Success]{}".format(self.rename_file.__name__,filename))
                    log_file.write('\n')
                return new_filename
            except FileExistsError as e:
                # for some reason, some files get copied and look like they don't exist in the file explorer. Look at them with 'll' they show. 
                # just deleting them in this case
                os.remove(filename)
                return new_filename 
            except Exception as e:
                with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                    log_file.write("[{} Failed]{} --- {}".format(self.rename_file.__name__,filename, str(e)))
                    log_file.write('\n')
                return False # Probably should so somthing different here? Not sure the best solution. 
        else:
            return False

    def group_by_extension(self):
        """
        Identifies extensions for all files in directories and sub-directories, \n
        groups similar file types together using the extension as the new directory name.\n
        Files are moved from the original directory to the new directory.
        """
        destination_dir = os.path.join(self.output_dir, "Grouped")
        
        if not os.path.exists(destination_dir):
            os.mkdir(destination_dir) 
        
        extensions = set()
        filename_gen = self.generate_filenames(self.original_dir)
        print("\nGrouping by file type. . .\n")
        for og_filepath in tqdm(filename_gen, total=self.estimated_files): 
            extension = Path(og_filepath).suffix.lower().replace(".", "") #Isolate the extension
            file_basename = os.path.split(og_filepath)[1]
            file_basename_no_extension = file_basename.split(".")[0]
            extension_dir = os.path.join(self.original_dir, extension) #Directory with the extension name, remove the '.'
            
            # Add new extensions to the set and create the directory if needed
            if extension not in extensions:
                extensions.add(extension) 
                if not os.path.isdir(extension_dir):
                    os.mkdir(extension_dir) 
            
            #Move to the group directory
            new_filepath = os.path.join(destination_dir, extension_dir, file_basename)
            try:
                if os.path.isfile(new_filepath): # Handle duplicate files by appending a unique value to the filename
                    now = int(time() * 1000) 
                    new_file_basename = file_basename_no_extension + "_" + str(now) + "." + extension
                    new_filepath = os.path.join(destination_dir, extension_dir, new_file_basename)
                move(og_filepath, new_filepath)
                with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                    log_file.write("[{} Success]{}".format(self.group_by_extension.__name__,og_filepath))
                    log_file.write('\n')
            except Exception as e:
                with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                    log_file.write("[{} Failed]{}".format(self.group_by_extension.__name__,str(e)))
                    log_file.write('\n')

    def organize(self, search_dir, filename, error=False, supported=True):
        """
        Method to move files to a new directory while retaining any sub directories.
        """
        if error is True:
            dest_dir = self.error_dir
        else:
            dest_dir = self.processed_dir
        if supported is False:
            dest_dir = self.unsupported_dir
        # Building out the directory structure. The goal is to perserve the relevant directories and save into a new directory. 
        og_dir = os.path.split(filename)[0] # Original file directory
        if og_dir == search_dir: # Condition for when we are searching the base directory
            og_dir_base = og_dir.replace(search_dir, "") # Directory path excluding the path we started from
        else: # Condition for when we are searching sub directories
            og_dir_base = og_dir.replace(search_dir + "\\", "") # Directory path excluding the path we started from
        og_file_basename = os.path.split(filename)[1] # Original file name
        partial_new_filepath = os.path.join(og_dir_base, og_file_basename) # New file path without the working directory
        full_new_dirpath = os.path.join(dest_dir, og_dir_base) # Combining the destination directory and the relevant directory structure.
        new_file = os.path.join(dest_dir, partial_new_filepath) # Bringing the components together

        # Move the file to the new location
        if not os.path.exists(full_new_dirpath):
            os.makedirs(full_new_dirpath)
        try:
            move(filename, new_file)
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Success]{}".format(self.organize.__name__,filename))
                log_file.write('\n')
            return True
        except Exception as e:
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Failure]{} --- {}".format(self.organize.__name__,filename, str(e)))
                log_file.write('\n')
            return False

    def generate_filenames(self, directory):
        """
        Simple function to generate filenames based on the files within the provided directory
        """
        filename_gen = glob.iglob(directory + '/**/*.*', recursive=True) #Generator for the filenames, recursively searches for all files in the directory.
        return filename_gen

    def get_file_magic(self, filename):
        """
        Detects the file type using a Python wrapper for libmagic
        """
        try:
            filetype = magic.from_file(filename)
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Success]{}".format(self.get_file_magic.__name__,filename))
                log_file.write('\n')
            return filetype
        except Exception as e:
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Failed]{} --- {}".format(self.get_file_magic.__name__,filename, str(e)))
                log_file.write('\n')
            return False

    def get_filetype_stats(self, directory):
        """
        Provides statistics for the total amount of files and each detected mimetype.
        Outputs results to a json file. 
        """
        file_type_stats = {'total':0} # Dictionary for storing stats
        filename_gen = self.generate_filenames(directory) # Files to process
        for filename in tqdm(filename_gen, total=self.estimated_files):
            try:
                # mimetype = magic.from_file(filename, mime=True).split("/")[0] # Detects mime type rather then file type, ommits data after the '/' to keep it more generic
                mimetype = magic.from_file(filename, mime=True) # Detects mime type rather then file type
                file_type_stats['total'] += 1
                if mimetype not in file_type_stats:
                    file_type_stats[mimetype] = 1
                else:
                    file_type_stats[mimetype] += 1
            except Exception as e:
                with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                    log_file.write("[{} Failed]{} --- {}".format(self.get_filetype_stats.__name__,filename, str(e)))
                    log_file.write('\n')
        
        try:
            stats_file = os.path.join(self.output_dir, "stats.json")
            with open(stats_file, 'w') as file:
                json.dump(file_type_stats, file, indent=4) # Output in a pretty print format
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Success]".format(self.get_filetype_stats.__name__))
                log_file.write('\n')
            return file_type_stats
        except Exception as e:
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Failed]{}".format(self.get_filetype_stats.__name__, str(e)))
                log_file.write('\n')
            return False
        
    def uncompress_tar(self, filename, dest_dir):
        """
        Uncompresses the file and saves to the destination directory
        """
        try:
            tar = tarfile.open(filename)
            members = tar.getmembers() # Contents of the compressed file
            tar.extractall(members=tqdm(members, desc=f"Uncompressing {filename}"), path=dest_dir) # Uncompress and use a progress bar
            tar.close()
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Success]{}".format(self.uncompress_tar.__name__,filename))
                log_file.write('\n')
            return True
        except Exception as e:
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Failed]{} --- {}".format(self.uncompress_tar.__name__,filename, str(e)))
                log_file.write('\n')
            return False

    def uncompress_tar_flevel(self, directory):
        """
        Uncompresses all files in the first level of the directory and saves to the destination directory
        """
        first_level_files = os.listdir(directory)
        for f in tqdm(first_level_files, desc="Checking for Compressed Files"):
            filename = os.path.join(directory, f)
            filetype = self.get_file_magic(filename)
            if filetype:
                if ".tar.gz" in filename.lower() or "gzip compressed data" in filetype.lower(): # Experienced some issues with python-magic classifying gzip files, using the less reliable file name as well. 
                    print(f"\nInspecting compressed file --- {filename}\n")
                    if self.uncompress_tar(filename, directory): # Uncompress file and save in the same directory
                        self.organize(directory, filename)
                    else:
                        self.organize(directory, filename, error=True)

    def cleanup_directories(self, directory):
        """
        Recursively deletes all empty directories. 
        Will continue to loop over the directories until no empties are found. 
        This allows for multiple levels of empty and nested directories to be deleted. 
        """
        while True:
            # Build set of empty directories
            empty_dirs = set()
            for (path,dirs,files) in os.walk(directory):
                if len(os.listdir(path)) == 0:
                    empty_dirs.add(path)
            
            # Exit when there are no empty directories
            if not empty_dirs:
                with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                    log_file.write("[{} Success]{}".format(self.cleanup_directories.__name__,directory))
                    log_file.write('\n')
                return True

            # Try and remove directories, exit if there are problems to avoid going into a loop
            else:
                for d in empty_dirs:
                    try:
                        os.rmdir(d)
                        with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                            log_file.write("[{} Success]{}".format(self.cleanup_directories.__name__,directory))
                            log_file.write('\n')
                        return True
                    except Exception as e:
                        with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                            log_file.write("[{} Failed]{} --- {}".format(self.cleanup_directories.__name__,directory, str(e)))
                            log_file.write('\n')
                        return False
    
    ########### Search methods, should probably be moved to their own class. 
    def _search_plaintext(self, filename):
        """
        Searches a provided text file using the keywords list included in the class. 
        Writes findings to a file, using the keyword as the filename. 
        """
        try:
            with open(filename, encoding='utf-8') as file: 
                for line in file:
                    for keyword in self.keywords:
                        if re.search(keyword, line, re.IGNORECASE):
                            keyword_result_file = os.path.join(self.results_dir, keyword + ".txt")
                            with open(keyword_result_file, 'a') as krs:
                                # file_basename = os.path.split(filename)[1] #Only the filename
                                result = filename + "---" + line
                                krs.write(result)
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Success]{}".format(self._search_plaintext.__name__,filename))
                log_file.write('\n')
            return True

        except Exception as e:
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Failure]{} --- {}".format(self._search_plaintext.__name__,filename, str(e)))
                log_file.write('\n')
            return False
    
    def _search_excel(self, filename):
        """
        Searches a modern excel file using the keywords list included in the class. 
        Writes findings to a file, using the keyword as the filename. 
        openpyxl is a Python library to read/write Excel 2010 xlsx/xlsm/xltx/xltm files.
        """
        try:
            wb = openpyxl.load_workbook(filename)
            sheets = wb.sheetnames
            
            for sheet in sheets:
                ws = wb[sheet]
                for row_cells in ws.iter_rows(max_row=10000): #setting a max row in case the file, some excel files can have seemingly endless rows even though they are blank
                    for cell in row_cells:
                        if cell.value:
                            for keyword in self.keywords:
                                if re.search(keyword, str(cell.value), re.IGNORECASE):
                                    keyword_result_file = os.path.join(self.results_dir, keyword + ".txt")
                                    with open(keyword_result_file, 'a') as krs:
                                        # file_basename = os.path.split(filename)[1] #Only the filename
                                        result = filename + "---" + cell.value
                                        krs.write(result)
                                        krs.write("\n")
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Success]{}".format(self._search_excel.__name__,filename))
                log_file.write('\n')
            return True

        except Exception as e:
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Failure]{} --- {}".format(self._search_excel.__name__,filename, str(e)))
                log_file.write('\n')
            return False

    def _search_excel_old_format(self, filename):
        """
        Searches a legacy excel file using the keywords list included in the class. 
        Writes findings to a file, using the keyword as the filename. 
        Assumes files are XLS format. 
        """
        try:
            wb = xlrd.open_workbook(filename)
            sheets = wb.sheet_names()
            for sheet_name in sheets:
                ws = wb.sheet_by_name(sheet_name)
                for row_idx in range(0, ws.nrows):    # Iterate through rows
                    for col_idx in range(0, ws.ncols):  # Iterate through columns
                        cell = ws.cell(row_idx, col_idx)  # Get cell object by row, col
                        if cell.value: # Ignore empty cells
                            for keyword in self.keywords:
                                if re.search(keyword, str(cell.value), re.IGNORECASE):
                                    keyword_result_file = os.path.join(self.results_dir, keyword + ".txt")
                                    with open(keyword_result_file, 'a') as krs:
                                        # file_basename = os.path.split(filename)[1] #Only the filename
                                        result = filename + "---" + cell.value
                                        krs.write(result)
                                        krs.write("\n")
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Success]{}".format(self._search_excel_old_format.__name__,filename))
                log_file.write('\n')
            return True
        except Exception as e:
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Failure]{} --- {}".format(self._search_excel_old_format.__name__,filename, str(e)))
                log_file.write('\n')
            return False

    def _search_word_docx(self, filename):
        """
        Searches a provided word document using the keywords list included in the class. 
        Writes findings to a file, using the keyword as the filename. 
        """
        try:
            # open connection to Word Document
            doc = docx.Document(filename)
        
            # read in each paragraph in file
            for paragraph in doc.paragraphs:
                for keyword in self.keywords:
                    if re.search(keyword, paragraph.text, re.IGNORECASE):
                        keyword_result_file = os.path.join(self.results_dir, keyword + ".txt")
                        with open(keyword_result_file, 'a') as krs:
                            # file_basename = os.path.split(filename)[1] #Only the filename
                            result = filename + "---" + paragraph.text
                            krs.write(result)
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Success]{}".format(self._search_word_docx.__name__,filename))
                log_file.write('\n')
            return True
        except Exception as e:
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Failure]{} --- {}".format(self._search_word_docx.__name__,filename, str(e)))
                log_file.write('\n')
            return False

    def _search_pdf(self, filename):
        try:
            with open(filename,'rb') as pdf_file:
                read_pdf = PyPDF2.PdfFileReader(pdf_file, strict=False) #Read and supress warnings
                if read_pdf.isEncrypted:
                    with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                        log_file.write("[{}]{} --- {}".format(self._search_pdf.__name__,filename, "File is encrypted, unable to parse."))
                        log_file.write('\n')
                    return False
                number_of_pages = read_pdf.getNumPages()
                for page_number in range(number_of_pages):
                    page = read_pdf.getPage(page_number)
                    page_content = page.extractText()
                    for keyword in self.keywords:
                        if re.search(keyword, page_content, re.IGNORECASE):
                            keyword_result_file = os.path.join(self.results_dir, keyword + ".txt")
                            with open(keyword_result_file, 'a') as krs:
                                # file_basename = os.path.split(filename)[1] #Only the filename
                                result = filename + "---" + page_content
                                krs.write(result)
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Success]{}".format(self._search_pdf.__name__,filename))
                log_file.write('\n')
            return True

        except Exception as e:
            with open(self.log_file, 'a', encoding=self.system_encoding) as log_file:
                log_file.write("[{} Failure]{} --- {}".format(self._search_pdf.__name__,filename, str(e)))
                log_file.write('\n')
            return False

    def _search_rich_text(self, filename):
        """UNDER DEVELOPMENT"""
        myfile = ""
        pattern = "{\\*?\\\\.+(;})|\\s?\\\\[A-Za-z0-9]+|\\s?{\\s?\\\\[A-Za-z0-9]+\\s?|\\s?}\\s?"
        with open(myfile) as file:
            for line in file:
                newline = re.sub(pattern, "", line)
                print(newline)
                # with open(self.log_file, 'a') as logfile:
                #     logfile.write(newline)

    def process_directory(self, directory):
        """ 
        Search files by trying to check the file type and use the appropriate method to parse. 
        """
        start_at_beginning = True # Used to restart the inner for loop as needed. 
        while start_at_beginning is True:
            start_at_beginning = False # Inner for loop must set this back to True if needed

            print("\nStarting from the beginning, generating a new list of file names. . .\n")
            filename_gen = self.generate_filenames(directory)

            print("\nDetecting file types and attempting to search. . .\n")
            for filename in tqdm(filename_gen, desc="Progress", total=self.estimated_files): #Can't find a good way to show progress while using a generator. Tried to use a list but ran into memory errors
                if self.rename_file(filename): # See if the file needed to be renamed. 
                    filename = self.rename_file(filename)
                filetype = self.get_file_magic(filename)
                if filetype:
                    # Compressed files
                    if ".tar.gz" in filename.lower() or "gzip compressed data" in filetype.lower(): # Experienced some issues with python-magic classifying gzip files, using the less reliable file name as well. 
                        print(f"\nInspecting compressed file --- {filename}\n")
                        if self.uncompress_tar(filename, directory): # Uncompress file and save in the same directory
                            self.organize(directory, filename)
                            start_at_beginning = True #When we uncompress a file, new files are expected. Time to restart with a new file generator. 
                            print("\nNew files have been uncompressed. Starting from the beginning. . .\n")
                            break
                        else:
                            self.organize(directory, filename, error=True)
                    # Text Files
                    elif "text" in filetype.lower():
                        if self._search_plaintext(filename):
                            self.organize(directory, filename)
                        else:
                            self.organize(directory, filename, error=True)
                    # Modern Excel Files
                    elif "Microsoft Excel 2007+" in filetype:
                        if self._search_excel(filename):
                            self.organize(directory, filename)
                        else:
                            self.organize(directory, filename, error=True)
                    # Old Excel Files
                    elif (".xls" in filename) and ("Composite Document File V2 Document, Little Endian" in filetype or "CDFV2 Microsoft Excel" in filetype):
                        if self._search_excel_old_format(filename):
                            self.organize(directory, filename)
                        else:
                            self.organize(directory, filename, error=True)
                    # Modern Word Files
                    elif "Microsoft Word 2007+" in filetype:
                        if self._search_word_docx(filename):
                            self.organize(directory, filename)
                        else:
                            self.organize(directory, filename, error=True)
                    # PDF Files
                    elif "PDF document" in filetype:
                        if self._search_pdf(filename):
                            self.organize(directory, filename)
                        else:
                            self.organize(directory, filename, error=True)
                    # Unsupported files
                    else:
                        self.organize(directory, filename, supported=False)

def main():
    # Variables & instantiation 
    working_dir = "" # Base directory we are working from
    original_dir = "" # Directory containing the files of interest
    keywords_file = "" # File containing the keywords which will be searched for
    estimated_files = None # You can put an integer here, best guess of how many files there are. Provides a more useful progress bar.

    fs = FileSearcher(working_dir, original_dir, keywords_file, estimated_files=estimated_files)

    # Create Required Directories
    fs.create_dirs()

    # Preliminary check for compressed files. Saves some processing time to do a first pass through the directory (non recursive) and uncompress files.
    # If additional compressed files are found during processing, they will be handled as well. However, the loop gets restarted to handle the newly uncompressed files.
    # fs.uncompress_tar_flevel(original_dir)

    # Get file statistics if needed, this is time consuming.
    # fs.get_filetype_stats(searchme)
    
    # Optionally, group the files by file exention. 
    # fs.group_by_extension()

    # Declare what we are searching
    searchme = original_dir # Likely either going to use the 'original_dir' or the Grouped file directory.
    # Search
    fs.process_directory(searchme)
    # Delete any empty directories
    fs.cleanup_directories(searchme)
    
if __name__ == "__main__":
    main()

